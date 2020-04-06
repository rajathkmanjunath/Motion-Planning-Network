import openpyxl

wb = openpyxl.load_workbook('../dataset/testfile.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
start_col_ind = sheet.max_row

data = [[1,2,3,4,5,6,7],[2,3,4,5,6,7,8],[3,4,5,6,7,8,9]]

for i in range(len(data)):
    for j in range(len(data[0])):
        sheet.cell(row=i+2+start_col_ind, column=j+1).value = data[i][j]

wb.save('../dataset/testfile.xlsx')